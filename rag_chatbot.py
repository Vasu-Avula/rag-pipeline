"""
RAG Pipeline Chatbot
====================
Full pipeline: Document Loading → Cleaning → Chunking → Embedding
               → Vector Store → Retrieval → LLM Answer

Supported file types: .txt, .pdf, .db (SQLite), .csv, .json, .md, .docx, .xlsx
"""

import os
import json
import sqlite3
import re
import math
import anthropic
from pathlib import Path


# ══════════════════════════════════════════════════════
# STEP 1 — DOCUMENT LOADING
# Reads every file from the given folder.
# Returns a list of {"source": filename, "text": raw_text}
# ══════════════════════════════════════════════════════

def load_txt(path: Path) -> str:
    """Load plain text / markdown files."""
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def load_pdf(path: Path) -> str:
    """Extract text from PDF using pypdf."""
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        return "\n".join(pages)
    except ImportError:
        print("  [warn] pypdf not installed. Run: pip install pypdf")
        return ""


def load_db(path: Path) -> str:
    """Extract all text content from SQLite database tables."""
    texts = []
    try:
        conn = sqlite3.connect(str(path))
        cursor = conn.cursor()
        # Get all table names
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = [row[0] for row in cursor.fetchall()]
        for table in tables:
            try:
                cursor.execute(f"SELECT * FROM {table} LIMIT 1000;")
                rows = cursor.fetchall()
                col_names = [desc[0] for desc in cursor.description]
                texts.append(f"[Table: {table}]")
                texts.append(" | ".join(col_names))  # Header row
                for row in rows:
                    texts.append(" | ".join(str(v) for v in row))
            except Exception as e:
                texts.append(f"[Could not read table {table}: {e}]")
        conn.close()
    except Exception as e:
        print(f"  [warn] Could not read DB {path.name}: {e}")
    return "\n".join(texts)


def load_csv(path: Path) -> str:
    """Load CSV as plain text (header + rows)."""
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def load_json(path: Path) -> str:
    """Load JSON file and convert to readable text."""
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            data = json.load(f)
        return json.dumps(data, indent=2)
    except Exception:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()


def load_docx(path: Path) -> str:
    """Extract text from .docx files."""
    try:
        import zipfile, xml.etree.ElementTree as ET
        texts = []
        with zipfile.ZipFile(str(path)) as z:
            with z.open("word/document.xml") as xml_file:
                tree = ET.parse(xml_file)
                root = tree.getroot()
                ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
                for para in root.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p"):
                    runs = para.findall(".//w:t", ns)
                    line = "".join(r.text or "" for r in runs)
                    if line.strip():
                        texts.append(line)
        return "\n".join(texts)
    except Exception as e:
        print(f"  [warn] Could not read DOCX {path.name}: {e}")
        return ""


def load_xlsx(path: Path) -> str:
    """Extract text from all sheets in an Excel (.xlsx / .xls) file."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True)
        texts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            texts.append(f"[Sheet: {sheet_name}]")
            for row in ws.iter_rows(values_only=True):
                # Skip completely empty rows
                if any(cell is not None for cell in row):
                    texts.append(" | ".join(str(cell) if cell is not None else "" for cell in row))
        return "\n".join(texts)
    except ImportError:
        print("  [warn] openpyxl not installed. Run: pip install openpyxl")
        return ""
    except Exception as e:
        print(f"  [warn] Could not read XLSX {path.name}: {e}")
        return ""


# Map file extensions → loader functions
LOADERS = {
    ".txt":    load_txt,
    ".md":     load_txt,
    ".pdf":    load_pdf,
    ".db":     load_db,
    ".sqlite": load_db,
    ".csv":    load_csv,
    ".json":   load_json,
    ".docx":   load_docx,
    ".xlsx":   load_xlsx,
    ".xls":    load_xlsx,
}


def load_documents(folder: str) -> list[dict]:
    """
    Walk the folder and load all supported files.
    Returns: [{"source": "file.pdf", "text": "...content..."}, ...]
    """
    folder_path = Path(folder)
    if not folder_path.exists():
        raise FileNotFoundError(f"Folder not found: {folder}")

    documents = []
    files = list(folder_path.rglob("*"))  # Recursive — includes subfolders

    print(f"\n{'═'*55}")
    print(f"  STEP 1 — LOADING DOCUMENTS from: {folder}")
    print(f"{'═'*55}")

    for file_path in files:
        if not file_path.is_file():
            continue
        ext = file_path.suffix.lower()
        if ext not in LOADERS:
            print(f"  [skip] {file_path.name} (unsupported type: {ext})")
            continue

        print(f"  [load] {file_path.name}", end=" ... ")
        loader = LOADERS[ext]
        text = loader(file_path)

        if text.strip():
            documents.append({"source": file_path.name, "text": text})
            print(f"✓ ({len(text):,} chars)")
        else:
            print("empty, skipped")

    print(f"\n  ✓ Loaded {len(documents)} document(s)\n")
    return documents


# ══════════════════════════════════════════════════════
# STEP 2 — CLEANING
# Removes noise: extra whitespace, control chars, etc.
# ══════════════════════════════════════════════════════

def clean_text(text: str) -> str:
    """Normalize and clean raw extracted text."""
    # Remove null bytes and control characters (keep newlines/tabs)
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", text)
    # Collapse multiple blank lines into two
    text = re.sub(r"\n{3,}", "\n\n", text)
    # Collapse runs of spaces/tabs (but keep newlines)
    text = re.sub(r"[ \t]+", " ", text)
    # Strip leading/trailing whitespace per line
    lines = [line.strip() for line in text.splitlines()]
    return "\n".join(lines).strip()


def clean_documents(documents: list[dict]) -> list[dict]:
    print(f"{'═'*55}")
    print(f"  STEP 2 — CLEANING TEXT")
    print(f"{'═'*55}")
    cleaned = []
    for doc in documents:
        original_len = len(doc["text"])
        clean = clean_text(doc["text"])
        cleaned.append({"source": doc["source"], "text": clean})
        print(f"  {doc['source']}: {original_len:,} → {len(clean):,} chars")
    print(f"\n  ✓ Cleaning complete\n")
    return cleaned


# ══════════════════════════════════════════════════════
# STEP 3 — CHUNKING
# Splits documents into smaller overlapping pieces
# so each chunk fits in the LLM context window.
# ══════════════════════════════════════════════════════

def chunk_text(text: str, chunk_size: int = 500, overlap: int = 100) -> list[str]:
    """
    Split text into chunks of ~chunk_size words with overlap.
    Uses word-level splitting to avoid cutting mid-sentence.
    """
    words = text.split()
    chunks = []
    start = 0
    while start < len(words):
        end = start + chunk_size
        chunk = " ".join(words[start:end])
        chunks.append(chunk)
        start += chunk_size - overlap  # Slide window with overlap
    return chunks


def chunk_documents(documents: list[dict],
                    chunk_size: int = 500,
                    overlap: int = 100) -> list[dict]:
    """
    Chunk all documents.
    Returns: [{"source": "file.pdf", "chunk_id": 0, "text": "..."}, ...]
    """
    print(f"{'═'*55}")
    print(f"  STEP 3 — CHUNKING  (size={chunk_size} words, overlap={overlap})")
    print(f"{'═'*55}")

    all_chunks = []
    for doc in documents:
        chunks = chunk_text(doc["text"], chunk_size, overlap)
        for i, chunk in enumerate(chunks):
            all_chunks.append({
                "source":   doc["source"],
                "chunk_id": i,
                "text":     chunk,
            })
        print(f"  {doc['source']}: {len(chunks)} chunk(s)")

    print(f"\n  ✓ Total chunks: {len(all_chunks)}\n")
    return all_chunks


# ══════════════════════════════════════════════════════
# STEP 4 — EMBEDDING
# Converts each chunk into a numeric vector.
# We use a lightweight TF-IDF style embedding (no GPU,
# no API cost) so the app works with zero extra deps.
# Swap embed_chunks() for OpenAI/Cohere if you prefer.
# ══════════════════════════════════════════════════════

def tokenize(text: str) -> list[str]:
    """Simple word tokenizer — lowercases and removes punctuation."""
    return re.findall(r"[a-z0-9]+", text.lower())


def build_vocabulary(chunks: list[dict]) -> dict[str, int]:
    """Build a vocabulary index from all chunk tokens."""
    vocab = {}
    idx = 0
    for chunk in chunks:
        for token in set(tokenize(chunk["text"])):
            if token not in vocab:
                vocab[token] = idx
                idx += 1
    return vocab


def tfidf_vector(text: str, vocab: dict, idf: dict) -> list[float]:
    """
    Compute a TF-IDF vector for a piece of text.
    TF  = term frequency in this text
    IDF = precomputed inverse document frequency
    """
    tokens = tokenize(text)
    if not tokens:
        return [0.0] * len(vocab)

    tf = {}
    for t in tokens:
        tf[t] = tf.get(t, 0) + 1
    total = len(tokens)

    vec = [0.0] * len(vocab)
    for token, count in tf.items():
        if token in vocab:
            idx = vocab[token]
            vec[idx] = (count / total) * idf.get(token, 0.0)
    return vec


def compute_idf(chunks: list[dict], vocab: dict) -> dict[str, float]:
    """Compute inverse document frequency for each token."""
    N = len(chunks)
    df = {}  # document frequency
    for chunk in chunks:
        tokens = set(tokenize(chunk["text"]))
        for t in tokens:
            df[t] = df.get(t, 0) + 1
    idf = {}
    for token in vocab:
        idf[token] = math.log((N + 1) / (df.get(token, 0) + 1)) + 1
    return idf


def embed_chunks(chunks: list[dict]) -> tuple[list[dict], dict, dict]:
    """
    Embed all chunks using TF-IDF vectors.
    Returns (chunks_with_vectors, vocab, idf)
    """
    print(f"{'═'*55}")
    print(f"  STEP 4 — EMBEDDING  (TF-IDF, {len(chunks)} chunks)")
    print(f"{'═'*55}")

    vocab = build_vocabulary(chunks)
    idf   = compute_idf(chunks, vocab)

    embedded = []
    for chunk in chunks:
        vec = tfidf_vector(chunk["text"], vocab, idf)
        embedded.append({**chunk, "vector": vec})

    print(f"  ✓ Vocabulary size : {len(vocab):,} tokens")
    print(f"  ✓ Embedded chunks : {len(embedded)}")
    print()
    return embedded, vocab, idf


# ══════════════════════════════════════════════════════
# STEP 5 — VECTOR STORE (in-memory)
# Stores all embedded chunks and supports cosine
# similarity search to find the most relevant ones.
# ══════════════════════════════════════════════════════

def cosine_similarity(a: list[float], b: list[float]) -> float:
    """Compute cosine similarity between two vectors."""
    dot   = sum(x * y for x, y in zip(a, b))
    mag_a = math.sqrt(sum(x * x for x in a))
    mag_b = math.sqrt(sum(x * x for x in b))
    if mag_a == 0 or mag_b == 0:
        return 0.0
    return dot / (mag_a * mag_b)


class VectorStore:
    """Simple in-memory vector store with cosine similarity search."""

    def __init__(self):
        self.chunks = []  # List of chunk dicts with "vector" field

    def add(self, embedded_chunks: list[dict]):
        self.chunks = embedded_chunks
        print(f"{'═'*55}")
        print(f"  STEP 5 — VECTOR STORE")
        print(f"{'═'*55}")
        print(f"  ✓ Stored {len(self.chunks)} vectors in memory\n")

    def search(self, query_vector: list[float], top_k: int = 5) -> list[dict]:
        """Return top_k most similar chunks to the query vector."""
        scored = []
        for chunk in self.chunks:
            score = cosine_similarity(query_vector, chunk["vector"])
            scored.append({**chunk, "score": score})
        scored.sort(key=lambda x: x["score"], reverse=True)
        return scored[:top_k]


# ══════════════════════════════════════════════════════
# STEP 6 — RETRIEVAL
# Embeds the user's query and finds the most
# relevant chunks from the vector store.
# ══════════════════════════════════════════════════════

def retrieve(query: str,
             store: VectorStore,
             vocab: dict,
             idf: dict,
             top_k: int = 5) -> list[dict]:
    """
    Embed the query and retrieve the top_k most relevant chunks.
    """
    query_vec = tfidf_vector(query, vocab, idf)
    results   = store.search(query_vec, top_k=top_k)
    return results


# ══════════════════════════════════════════════════════
# STEP 7 — BUILD CONTEXT + CALL LLM
# Assembles retrieved chunks into a context block
# and sends it to Claude with the user's question.
# ══════════════════════════════════════════════════════

SYSTEM_PROMPT = """You are a helpful assistant that answers questions based only on the provided context documents.

RULES:
- Answer using ONLY the information in the context below.
- If the context does not contain enough information to answer, say so clearly.
- Always cite which document(s) your answer comes from.
- Be concise and accurate."""


def build_context(retrieved_chunks: list[dict]) -> str:
    """Format retrieved chunks into a readable context block."""
    parts = []
    for i, chunk in enumerate(retrieved_chunks, 1):
        parts.append(
            f"[Source {i}: {chunk['source']} | chunk {chunk['chunk_id']} | "
            f"relevance: {chunk['score']:.3f}]\n{chunk['text']}"
        )
    return "\n\n---\n\n".join(parts)


def ask_llm(client: anthropic.Anthropic,
            query: str,
            context: str,
            history: list[dict]) -> str:
    """
    Call Claude with the retrieved context and the user's question.
    Maintains conversation history for multi-turn chat.
    """
    # Build the user message with context injected
    user_message = f"""Context documents:
{context}

---

Question: {query}"""

    # Append this turn to history
    history.append({"role": "user", "content": user_message})

    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=1024,
        system=SYSTEM_PROMPT,
        messages=history,
    )

    answer = response.content[0].text

    # Append assistant reply to history for next turn
    history.append({"role": "assistant", "content": answer})

    return answer


# ══════════════════════════════════════════════════════
# MAIN — Full Pipeline + Chat Loop
# ══════════════════════════════════════════════════════

def build_pipeline(folder: str) -> tuple[VectorStore, dict, dict]:
    """
    Run the full ingestion pipeline:
    Load → Clean → Chunk → Embed → Store
    Returns the vector store and embedding metadata.
    """
    documents     = load_documents(folder)
    if not documents:
        raise ValueError(f"No supported documents found in: {folder}")

    cleaned       = clean_documents(documents)
    chunks        = chunk_documents(cleaned, chunk_size=500, overlap=100)
    embedded, vocab, idf = embed_chunks(chunks)

    store = VectorStore()
    store.add(embedded)

    return store, vocab, idf


def chat_loop(client: anthropic.Anthropic,
              store: VectorStore,
              vocab: dict,
              idf: dict):
    """Interactive chat loop — ask questions about your documents."""
    history = []  # Conversation history for multi-turn context

    print(f"{'═'*55}")
    print(f"  STEP 6+7 — RETRIEVAL + LLM  (RAG Chat Ready!)")
    print(f"{'═'*55}")
    print("  Type your question and press Enter.")
    print("  Type 'quit' or 'exit' to stop.\n")

    while True:
        try:
            query = input("You: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nGoodbye!")
            break

        if not query:
            continue
        if query.lower() in ("quit", "exit", "q"):
            print("Goodbye!")
            break

        # Retrieve relevant chunks
        print("\n  [retrieving relevant chunks...]")
        retrieved = retrieve(query, store, vocab, idf, top_k=5)

        # Show which sources were found
        sources = list(dict.fromkeys(c["source"] for c in retrieved))
        print(f"  [sources: {', '.join(sources)}]")

        # Build context from retrieved chunks
        context = build_context(retrieved)

        # Call the LLM
        print("  [calling LLM...]\n")
        try:
            answer = ask_llm(client, query, context, history)
            print(f"Assistant: {answer}\n")
        except anthropic.BadRequestError as e:
            if "credit" in str(e).lower():
                print("  [error] Anthropic account has no credits. Add credits at console.anthropic.com\n")
            else:
                print(f"  [error] {e}\n")
        except Exception as e:
            print(f"  [error] {e}\n")


def main():
    # ── Read API key ──────────────────────────────
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise EnvironmentError(
            "ANTHROPIC_API_KEY environment variable is not set.\n"
            "Windows PowerShell: $env:ANTHROPIC_API_KEY='your-key-here'"
        )

    # ── Choose documents folder ───────────────────
    folder = os.environ.get("DOCS_FOLDER", "documents")
    print(f"\n{'═'*55}")
    print(f"  RAG PIPELINE CHATBOT")
    print(f"{'═'*55}")
    print(f"  Documents folder : {folder}")
    print(f"  Supported types  : txt, pdf, db, csv, json, md, docx")

    # ── Run ingestion pipeline ────────────────────
    store, vocab, idf = build_pipeline(folder)

    # ── Start chat ────────────────────────────────
    client = anthropic.Anthropic(api_key=api_key)
    chat_loop(client, store, vocab, idf)


if __name__ == "__main__":
    main()
